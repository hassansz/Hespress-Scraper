import requests
import aiohttp
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook  # Updated this line
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import time
from aiohttp import ClientSession
import nest_asyncio

# Allow nested event loops (for Google Colab)
nest_asyncio.apply()

# Constants
MAX_PAGES = 2500000
OUTPUT_FILE = 'hespress.xlsx'
CONCURRENT_REQUESTS = 10
DEBUG = True  # Enable detailed debugging

# Headers
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# Create or load Excel workbook
if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)  # Load existing workbook
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hespress Data"
    # Add headers
    headers_row = ['Title', 'Category', 'Date', 'Content', 'Link', 'Image', 'Tags', 'Number of Comments', 'Comments']
    ws.append(headers_row)
    for col in range(1, 10):  # Updated to 10 columns
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

def scrape_page(page_num):
    try:
        print(f'Filtering data from page {page_num} :')
        url = f'https://www.hespress.com/?action=ajax_listing&paged={page_num}&tq=MjAyMi0wNC0wNiAwMDowNTowMA%3D%3D&all_listing=1'
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Raise an exception for bad status codes

        soup = BeautifulSoup(response.text, 'lxml')
        posts = soup.find_all('div', class_='overlay card')

        if not posts:
            print(f"No posts found on page {page_num}. Might be the end.")
            return False

        for post in posts:
            try:
                print(f'[+] Scraping post from page {page_num}')
                # Extract basic fields
                post_title = post.find('h3', class_='card-title').text.strip()
                post_category = post.find('span', class_='cat').text.strip()
                post_img = post.find('img', class_='wp-post-image')['src']
                post_link = post.find('a', class_='stretched-link')['href']
                post_date = post.find('small', class_='text-muted time').text.strip()

                # Get post content and additional details
                post_response = requests.get(post_link, headers=headers)
                post_response.raise_for_status()
                post_soup = BeautifulSoup(post_response.text, 'lxml')

                # Content
                content_div = post_soup.find('div', class_='article-content')
                post_content = ''
                if content_div:
                    paragraphs = content_div.find_all('p')
                    post_content = '\n'.join(p.text.strip() for p in paragraphs)

                # Tags
                tags_section = post_soup.find('section', class_='box-tags')
                post_tags = "No Tags"
                if tags_section:
                    tags = tags_section.find_all('a', class_='tag_post_tag')
                    post_tags = ", ".join(tag.get_text().strip() for tag in tags) if tags else "No Tags"

                # Number of Comments
                comment_count_tag = post_soup.find('span', class_='comments-count-number')
                post_comment_count = comment_count_tag.get_text().strip() if comment_count_tag else "0"

                # Comments
                comments_section = post_soup.find('div', class_='comments')
                post_comments = []
                if comments_section:
                    comment_list = comments_section.find('ul', class_='comment-list')
                    if comment_list:
                        comments = comment_list.find_all('li', class_='comment')
                        for comment in comments:
                            comment_body = comment.find('div', class_='comment-body')
                            if comment_body:
                                name_tag = comment_body.find('span', class_='comment-author') or comment_body.find('div', class_='comment-author')
                                commenter_name = name_tag.get_text().strip() if name_tag else "Anonymous"
                                comment_text = comment_body.get_text().strip()
                                if name_tag:
                                    comment_text = comment_text.replace(commenter_name, '').strip()
                                post_comments.append(f"{commenter_name}: {comment_text}")
                post_comments_str = "\n".join(post_comments) if post_comments else "No Comments"

                # Append to Excel
                ws.append([post_title, post_category, post_date, post_content, post_link, post_img, post_tags, post_comment_count, post_comments_str])
                print('[+] Scraping post successfully')

            except Exception as e:
                print(f"Error scraping individual post: {e}")
                continue

        # Save progress after each page
        wb.save(OUTPUT_FILE)
        return True

    except Exception as e:
        print(f"Error scraping page {page_num}: {e}")
        return False

## Scrape up to MAX_PAGES
for page in range(1, MAX_PAGES + 1):
    if not scrape_page(page):
        print(f"Stopping scrape at page {page} due to failure or no more posts.")
        break
    time.sleep(2)  # Increased delay to avoid rate limiting

print(f"Scraping completed. Data saved to {OUTPUT_FILE}")

# Download the file in Google Colab
    from google.colab import files
    files.download(OUTPUT_FILE)
